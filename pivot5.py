import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl.styles import Font
from openpyxl.styles import Border, Side


# === Step 1: Load file and clean ===
file_path = "8 July 2025 Archer Toxic sharing.xlsx"
df = pd.read_excel(file_path, sheet_name="Archer Search Report (2)")
df.columns = df.columns.str.strip()

# === Step 2: Filter relevant FLT + Local rows ===
flt_local_df = df[
    (df["Current Status"] == "Forward Looking Toxic") &
    (df["IT Component Type"] == "Regional/Local")
]

# === Step 3: Ensure all OEs are present ===
all_oe_list = [
    "Allianz China - Holding", "Allianz China - P&C", "Allianz Indonesia",
    "Allianz Malaysia", "Allianz Philippine - L&H", "Allianz Singapore",
    "Allianz Sri Lanka", "Allianz Taiwan - Life", "Allianz Thailand"
]

flt_local_df = flt_local_df[
    ["Allianz OE Name", "IT Component Name", "Release", "Toxic from Date", "Number of IT Assets"]
]

flt_local_df["Number of IT Assets"] = flt_local_df["Number of IT Assets"].fillna(0)
flt_local_df["Toxic from Date"] = pd.to_datetime(flt_local_df["Toxic from Date"], errors='coerce')

# Create full list of component-detail triplets
component_details = flt_local_df[["IT Component Name", "Release", "Toxic from Date"]].drop_duplicates()

# Create pivot-style mapping for each OE
pivot_dict = {}
for oe in all_oe_list:
    row_data = {}
    for _, row in component_details.iterrows():
        key = (row["IT Component Name"], row["Release"], row["Toxic from Date"])
        val = flt_local_df[
            (flt_local_df["Allianz OE Name"] == oe) &
            (flt_local_df["IT Component Name"] == key[0]) &
            (flt_local_df["Release"] == key[1]) &
            (flt_local_df["Toxic from Date"] == key[2])
        ]["Number of IT Assets"].sum()
        row_data[key] = val
    pivot_dict[oe] = row_data

# === Step 4: Build Excel workbook ===
wb = Workbook()
ws = wb.active
ws.title = "FLT Local Details"

# === Multi-row header ===
# === Title Row (A1:C1) ===
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
ws.cell(row=1, column=1, value="Number of IT Assets")
ws.cell(row=6, column=1, value="Current Status")
ws.cell(row=6, column=2, value="IT Component Type")
ws.cell(row=6, column=3, value="Allianz OE Name")
ws.cell(row=3, column=4, value="IT Component Name")  		
ws.cell(row=3, column=5, value="Release")
ws.cell(row=3, column=6, value="Toxic from Date")


start_col = 4
component_keys = sorted(
    component_details.itertuples(index=False),
    key=lambda x: (x[0], str(x[1]), str(x[2]))
)   

for idx, key in enumerate(component_keys):
    col = start_col + idx
    comp_name, release, toxic_date = key
    ws.cell(row=4, column=col, value=comp_name)
    ws.cell(row=5, column=col, value=release)
    ws.cell(row=6, column=col, value=toxic_date.strftime("%m/%d/%Y") if pd.notna(toxic_date) else "")

# === Component Detail Header Loop ===
for idx, key in enumerate(component_keys):
    col = start_col + idx
    comp_name, release, toxic_date = key
    ws.cell(row=4, column=col, value=comp_name)
    ws.cell(row=5, column=col, value=release)
    ws.cell(row=6, column=col, value=toxic_date.strftime("%m/%d/%Y") if pd.notna(toxic_date) else "")

# === Merge Row 4 cells for consecutive duplicate IT Component Names ===
merge_start = start_col
prev_name = component_keys[0][0]

for idx, key in enumerate(component_keys[1:], start=1):
    col = start_col + idx
    curr_name = key[0]

    if curr_name != prev_name:
        if col - 1 > merge_start:
            ws.merge_cells(start_row=4, start_column=merge_start, end_row=4, end_column=col - 1)
        merge_start = col
    prev_name = curr_name

# Merge the final group (if any)
last_col = start_col + len(component_keys) - 1
if last_col > merge_start:
    ws.merge_cells(start_row=4, start_column=merge_start, end_row=4, end_column=last_col)


# Add headers for total columns
ws.cell(row=3, column=start_col + len(component_keys), value="Grand Total")
ws.cell(row=3, column=start_col + len(component_keys) + 1, value="Grand Total 2025")
ws.cell(row=3, column=start_col + len(component_keys) + 1).fill = openpyxl.styles.PatternFill(
    start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
)


# === Data rows ===
totals_by_col = [0] * len(component_keys)
totals_2025_by_row = []
grand_totals_by_row = []

for i, oe in enumerate(all_oe_list, start=7):
    ws.cell(row=i, column=1, value="Forward Looking Toxic")
    ws.cell(row=i, column=2, value="Regional/Local")
    ws.cell(row=i, column=3, value=oe)
    
    row_sum = 0
    row_sum_2025 = 0
    for j, key in enumerate(component_keys):
        val = pivot_dict[oe].get(key, 0)
        col = start_col + j
        ws.cell(row=i, column=col, value=val)
        row_sum += val
        totals_by_col[j] += val

        if pd.notna(key[2]) and key[2].year == 2025:
            row_sum_2025 += val

    grand_totals_by_row.append(row_sum)
    totals_2025_by_row.append(row_sum_2025)
    ws.cell(row=i, column=start_col + len(component_keys), value=row_sum)
    ws.cell(row=i, column=start_col + len(component_keys) + 1, value=row_sum_2025)

# === Total row ===
total_row_index = 6 + len(all_oe_list)
ws.cell(row=total_row_index, column=1, value="Grand Total")
ws.cell(row=total_row_index, column=2, value="")
ws.cell(row=total_row_index, column=3, value="")

# Fill in totals by column
for j, total in enumerate(totals_by_col):
    col = start_col + j
    ws.cell(row=total_row_index, column=col, value=total)

# Grand column total & 2025 column total
ws.cell(row=total_row_index, column=start_col + len(component_keys), value=sum(grand_totals_by_row))
ws.cell(row=total_row_index, column=start_col + len(component_keys) + 1, value=sum(totals_2025_by_row))



# Bold all numbers in total row
for col in range(4, start_col + len(component_keys) + 2):
    ws.cell(row=total_row_index, column=col).font = Font(bold=True)

# Bold "Grand Total 2025" label
ws.cell(row=3, column=start_col + len(component_keys) + 1).font = Font(bold=True)


for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2  # add padding

ws.freeze_panes = "D7"  # freeze everything left of column D and above row 7

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Apply border to every non-empty cell
for row in ws.iter_rows(
    min_row=3, 
    max_row=total_row_index,
    min_col=1,
    max_col=start_col + len(component_keys) + 1
):
    for cell in row:
        if cell.value is not None:
            cell.border = thin_border

# === Save file ===
wb.save("FLT_Local_Details_Final.xlsx")
# Auto-adjust column widths based on max content in each column
