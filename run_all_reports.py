from openpyxl import Workbook, load_workbook
from flt_pvt import generate_flt_pvt_sheet
from toxic_pvt import generate_toxic_pvt_sheet
from Group_FLT_Details import generate_group_flt_details
from Group_Toxic_Details import generate_group_toxic_details
from Local_FLT_Details import generate_local_flt_details
from Local_Toxic_Details import generate_local_toxic_details

OUTPUT_FILE = "8 July 2025 Archer Toxic sharing.xlsx"

try:
    wb = load_workbook(OUTPUT_FILE)
except FileNotFoundError:
    wb = Workbook()
    # Remove the default sheet if present
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

# Call all your sheet generators
generate_flt_pvt_sheet(wb)
generate_toxic_pvt_sheet(wb)
generate_group_flt_details(wb)
generate_group_toxic_details(wb)
generate_local_flt_details(wb)
generate_local_toxic_details(wb)

# Save once at the end
wb.save(OUTPUT_FILE)
print("All reports generated successfully!")
