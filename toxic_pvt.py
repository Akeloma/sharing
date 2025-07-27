import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import Font


# === Step 1: Load data ===
file_path = "8 July 2025 Archer Toxic sharing.xlsx"
df = pd.read_excel(file_path, sheet_name="Archer Search Report (2)")
df.columns = df.columns.str.strip()

# === Step 2: Filter for FLT only ===
flt_df = df[df["Current Status"] == "Toxic"]

# === Step 3: Group and Pivot ===
grouped = (
    flt_df.groupby(["Allianz OE Name", "IT Component Type"])["Number of IT Assets"]
    .sum()
    .unstack(fill_value=0)
    .reset_index()
)

grouped["Current Status"] = "Forward Looking Toxic"
grouped["Grand Total"] = grouped.get("Group", 0) + grouped.get("Regional/Local", 0)

# Rearranging columns to desired order
final_df = grouped[["Current Status", "Allianz OE Name", "Group", "Regional/Local", "Grand Total"]]

# === Step 4: Write to Excel ===
wb = Workbook()
ws = wb.active
ws.title = "Toxic pvt"

ws.cell(row=6, column=1, value="Sum of Number of IT")

# Write headers and data starting from row 7
for i, col_name in enumerate(final_df.columns, start=1):
    ws.cell(row=7, column=i, value=col_name)

for row_idx, row in final_df.iterrows():
    for col_idx, val in enumerate(row, start=1):
        ws.cell(row=8 + row_idx, column=col_idx, value=val)

# === Add Grand Total row ===
last_data_row = 8 + len(final_df)
ws.cell(row=last_data_row, column=1, value="Grand Total")
ws.cell(row=last_data_row, column=2, value="")  # Empty Allianz OE Name

bold_font = Font(bold=True)
for col in range(1, 6):
    ws.cell(row=last_data_row, column=col).font = bold_font

group_total = final_df["Group"].sum()
local_total = final_df["Regional/Local"].sum()
grand_total = final_df["Grand Total"].sum()

ws.cell(row=last_data_row, column=3, value=group_total)
ws.cell(row=last_data_row, column=4, value=local_total)
ws.cell(row=last_data_row, column=5, value=grand_total)


# Center align from column C onwards (3 to 5)
for row in ws.iter_rows(min_row=7, max_row=last_data_row, min_col=3, max_col=5):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

ws.sheet_view.showGridLines = False

thin = Side(style="thin")
border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
border_tb = Border(top=thin, bottom=thin)
border_lr = Border(left=thin, right=thin)

# Row 6 (top headers): top & bottom border
for cell in ws["6"]:
    col = cell.column
    if col <= 5:
        if col in [1, 2]:
            cell.border = border_all
        else:
            cell.border = border_tb

# Row 7 to data end (values): side borders only
for row in ws.iter_rows(min_row=7, max_row=last_data_row, min_col=3, max_col=5):
    for cell in row:
        cell.border = border_lr

# Add right border for the last header cell in row 6 (E6)
ws.cell(row=6, column=5).border = Border(top=thin, bottom=thin, right=thin)

# Current Status + OE Name: full borders
for row in ws.iter_rows(min_row=7, max_row=last_data_row, min_col=1, max_col=2):
    for cell in row:
        cell.border = border_lr

# Grand Total Row: top + bottom border
for col in range(1, 6):
    cell = ws.cell(row=last_data_row, column=col)
    if col == 5:
        cell.border = border_all
    elif col in [1, 2]:
        cell.border = border_all
    else:
        cell.border = border_tb

# Auto-fit column widths with padding
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 3  # Add padding of 3

# Save the file
wb.save("FLT_pvt_output.xlsx")
